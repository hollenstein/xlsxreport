import os

import pandas as pd
import pytest
import openpyxl
import xlsxwriter

import xlsxreport
from xlsxreport.writer import TableSectionWriter
from xlsxreport.template import ReportTemplate
from xlsxreport.compiler import prepare_table_sections


TESTDATA_DIRECTORY = os.path.join(os.path.dirname(__file__), "testdata")


@pytest.fixture()
def temp_excel_path(request, tmp_path):
    output_path = os.path.join(tmp_path, "output.xlsx")

    def teardown():
        if os.path.isfile(output_path):
            os.remove(output_path)

    request.addfinalizer(teardown)
    return output_path


class TestCorrectCreationOfFormattedExcelFile:
    def test_new_xlsx_report_implementation(self, temp_excel_path):
        template_path = os.path.join(TESTDATA_DIRECTORY, "mq_protein_template.yaml")
        mq_path = os.path.join(TESTDATA_DIRECTORY, "mq_proteinGroups.txt")
        table = pd.read_csv(mq_path, sep="\t")

        report_template = ReportTemplate.load(template_path)
        compiled_table_sections = prepare_table_sections(report_template, table)
        with xlsxwriter.Workbook(temp_excel_path) as workbook:
            worksheet = workbook.add_worksheet("Proteins")
            section_writer = TableSectionWriter(workbook)
            section_writer.write_sections(
                worksheet, compiled_table_sections, settings=report_template.settings
            )

        reference_excel_path = os.path.join(TESTDATA_DIRECTORY, "mq_proteinGroups.xlsx")
        with open(reference_excel_path, "rb") as f1, open(temp_excel_path, "rb") as f2:
            wb1 = openpyxl.load_workbook(f1)
            wb2 = openpyxl.load_workbook(f2)

            # Compare the number of sheets
            assert len(wb1.sheetnames) == len(wb2.sheetnames)

            # Compare the sheet names
            for sheet1, sheet2 in zip(wb1, wb2):
                assert sheet1.title == sheet2.title

            # Compare equal number of rows and columns
            assert len(list(sheet1.iter_rows())) == len(list(sheet2.iter_rows()))
            assert len(list(sheet1.iter_cols())) == len(list(sheet2.iter_cols()))

            # Compare the cell values and formatting
            for sheet1, sheet2 in zip(wb1, wb2):
                for row1, row2 in zip(sheet1.iter_rows(), sheet2.iter_rows()):
                    for cell1, cell2 in zip(row1, row2):
                        assert cell1.value == cell2.value
                        assert cell1._style == cell2._style
                        assert str(cell1.number_format) == str(cell2.number_format)
                        assert str(cell1.font) == str(cell2.font)
                        assert str(cell1.border) == str(cell2.border)
                        assert str(cell1.fill) == str(cell2.fill)
                        assert str(cell1.alignment) == str(cell2.alignment)

            # Compare the conditional formatting rules
            for sheet1, sheet2 in zip(wb1, wb2):
                for rule1, rule2 in zip(
                    sheet1.conditional_formatting._cf_rules,
                    sheet2.conditional_formatting._cf_rules,
                ):
                    assert rule1.cells == rule2.cells
                    for cfrule1, cfrul2 in zip(rule1.rules, rule2.rules):
                        assert str(cfrule1) == str(cfrul2)
