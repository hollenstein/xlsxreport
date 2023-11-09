from io import BytesIO
from unittest.mock import MagicMock, call

import openpyxl
import pandas as pd
import pytest
from xlsxwriter import Workbook as Workbook

import xlsxreport.compiler as compiler
import xlsxreport.excel_writer as writer


class ExcelWriteReadTestManager:
    """Helper class for writing an excel file with `xlsxwriter` to buffer and reading it
    from the buffer with `openpyxl`.

    Note that for writing rows and columns xlsxwriter is zero indexed whereas pyopenxl
    is one indexed.
    """

    def __init__(self):
        self.buffer = None
        self.loaded_workbook = None
        self.loaded_worksheet = None

    def __enter__(self):
        self.buffer = BytesIO()
        self.workbook = Workbook(self.buffer)
        self.worksheet_name = "Worksheet"
        self.worksheet = self.workbook.add_worksheet(self.worksheet_name)
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.workbook.close()
        self._load_excel_from_buffer()

    def _load_excel_from_buffer(self):
        if self.buffer is None:
            raise Exception("Buffer is not initialized")
        self.loaded_workbook = openpyxl.load_workbook(self.buffer)
        self.loaded_worksheet = self.loaded_workbook[self.worksheet_name]


@pytest.fixture()
def table_section() -> compiler.TableSection:
    section = compiler.TableSection(
        data=pd.DataFrame({"Column 1": [1, 2, 3], "Column 2": ["A", "B", "C"]}),
        column_widths={"Column 1": 10, "Column 2": 10},
    )
    return section


class TestTableSectionWriteSections:
    """Tests for TableSectionWriter.write_section():
    - Test that multiple sections are written with correct column and row positions
    - Test that supheader, header and row heights are set with correct values and on
    correct coordinates
    - Test that freeze panes is applied to the correct coordinates
    - Test that auto filter is applied to the correct coordinates
    """

    @pytest.fixture(autouse=True)
    def _init_table_sections(self):
        section_1 = compiler.TableSection(
            data=pd.DataFrame({"Column 1": [1, 2, 3], "Column 2": ["A", "B", "C"]}),
        )
        section_2 = compiler.TableSection(
            data=pd.DataFrame({"Column 3": [1, 2, 3], "Column 4": ["A", "B", "C"]}),
        )
        self.headers = list(section_1.headers) + list(section_2.headers)
        self.table_sections = [section_1, section_2]
        self.col_num = 4
        self.row_num = 3
        self.column_values = []
        for section in self.table_sections:
            self.column_values.extend([section.data[c].tolist() for c in section.data.columns])  # fmt: skip

    @pytest.mark.parametrize("start_row", [0, 2, 10])
    def test_all_section_headers_correctly_written_with_different_start_rows(self, start_row):  # fmt: skip
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer.write_sections(
                excel_manager.worksheet,
                self.table_sections,
                settings={"write_supheader": False},
                start_row=start_row,
            )
        sheet = excel_manager.loaded_worksheet
        assert [cell.value for cell in list(sheet.rows)[start_row]] == self.headers

    def test_all_section_values_are_written_to_the_correct_position(self):
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer.write_sections(
                excel_manager.worksheet,
                self.table_sections,
                settings={"write_supheader": False},
            )
        sheet_columns = list(excel_manager.loaded_worksheet.columns)
        for column_cells, column_values in zip(sheet_columns, self.column_values):
            value_cells = column_cells[1:]  # without header cell
            assert [cell.value for cell in value_cells] == column_values

    @pytest.mark.parametrize("start_col", [0, 2, 10])
    def test_start_column_correctly_applied(self, start_col):
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer.write_sections(
                excel_manager.worksheet, self.table_sections, start_column=start_col
            )
        sheet = excel_manager.loaded_worksheet
        for empty_column in list(sheet.columns)[:start_col]:
            assert all([cell.value is None for cell in empty_column])
        assert len(list(sheet.columns)) == self.col_num + start_col

    @pytest.mark.parametrize(
        "write_supheader, start_row", [(True, 0), (True, 2), (False, 0), (False, 2)]
    )
    def test_start_row_and_supheader_correctly_applied(
        self, write_supheader, start_row
    ):
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer.write_sections(
                excel_manager.worksheet,
                self.table_sections,
                settings={"write_supheader": write_supheader},
                start_row=start_row,
            )
        sheet_rows = list(excel_manager.loaded_worksheet.rows)
        header_row = start_row + int(write_supheader)
        assert [cell.value for cell in sheet_rows[header_row]] == self.headers


class TestIntegrationTableSectionWriteColumn:
    @pytest.fixture(autouse=True)
    def _init_section_writer_with_write_column(self):
        # Note that xlsxwriter is zero indexed whereas pyopenxl is one indexed
        row, column = 2, 2
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer._write_column(
                excel_manager.worksheet,
                row=row - 1,
                column=column - 1,
                header="Header",
                values=[1, 2, "3"],
                header_format={"bold": True, "bottom": 2},
                values_format={"align": "center", "num_format": "0.00"},
                conditional_format={"type": "2_color_scale"},
                column_width=10000,
            )
        self.worksheet = excel_manager.loaded_worksheet
        self.written_column = list(self.worksheet.columns)[1]

    def test_column_is_written_to_correct_position(self):
        columns = list(self.worksheet.columns)
        empty_col_position = 0
        written_col_position = 1
        assert all([cell.value is None for cell in columns[empty_col_position]])
        assert any([cell.value is None for cell in columns[written_col_position]])

    def test_correct_column_values_written(self):
        column_values = [cell.value for cell in self.written_column]
        assert column_values == [None, "Header", 1, 2, "3"]

    def test_header_format_is_applied(self):
        header_cell = self.worksheet.cell(row=2, column=2)
        assert header_cell.font.bold == True
        assert header_cell.border.bottom.style == "medium"

    def test_column_format_is_applied(self):
        column_cells = self.written_column[2:]  # without empty and header cell
        assert all([cell.alignment.horizontal == "center" for cell in column_cells])
        assert all([cell.number_format == "0.00" for cell in column_cells])

    def test_set_correct_column_width(self):
        # xlsxwriter sets column width in pixel units, whereas openpyxl sets width in
        # some kind of "unit" format. So we set the width to a high number and check if
        # the column width in "units" is smaller or bigger than 100.
        empty_col_width = self.worksheet.column_dimensions["A"].width
        written_col_width = self.worksheet.column_dimensions["B"].width
        assert empty_col_width < 100
        assert written_col_width > 100

    def test_conditional_format_applied_to_correct_area(self):
        conditional_format = list(self.worksheet.conditional_formatting)[0]
        conditional_format.cells.ranges[0].top == [(3, 2)]
        conditional_format.cells.ranges[0].bottom == [(5, 2)]


class TestTableSectionWriteColumn:
    @pytest.fixture(autouse=True)
    def _init_section_writer(self):
        self.worksheet_mock = MagicMock(name="worksheet_mock")
        self.section_writer = writer.TableSectionWriter(Workbook())

    @pytest.fixture(autouse=True)
    def _init_arguments_for_write_column(self):
        self.args = {
            "row": 1,
            "column": 1,
            "header": "Header name",
            "values": [1, 2, 3],
            "header_format": {"bold": True},
            "values_format": {"bold": False},
            "conditional_format": {"type": "cell"},
            "column_width": 10,
        }
        self.header_row = self.args["row"]
        self.data_row_start = self.args["row"] + 1
        self.data_row_end = self.args["row"] + len(self.args["values"])

    def test_write_column_called_with_correct_arguments_to_write_values(self):
        self.section_writer._write_column(self.worksheet_mock, **self.args)
        assert self.worksheet_mock.write_column.call_args == call(
            self.data_row_start,
            self.args["column"],
            self.args["values"],
            self.section_writer.get_xlsx_format(self.args["values_format"]),
        )

    def test_write_called_with_correct_arguments_to_write_header(self):
        self.section_writer._write_column(self.worksheet_mock, **self.args)
        assert self.worksheet_mock.write.call_args == call(
            self.header_row,
            self.args["column"],
            self.args["header"],
            self.section_writer.get_xlsx_format(self.args["header_format"]),
        )

    def test_set_column_pixels_called_with_correct_arguments(self):
        self.section_writer._write_column(self.worksheet_mock, **self.args)
        assert self.worksheet_mock.set_column_pixels.call_args == call(
            self.args["column"], self.args["column"], self.args["column_width"]
        )

    def test_conditional_format_called_with_correct_arguments(self):
        self.section_writer._write_column(self.worksheet_mock, **self.args)

        assert self.worksheet_mock.conditional_format.call_args == call(
            self.data_row_start,
            self.args["column"],
            self.data_row_end,
            self.args["column"],
            self.args["conditional_format"],
        )

    def test_conditional_format_not_called_when_conditionaL_format_is_empty(self):
        self.args["conditional_format"] = {}
        self.section_writer._write_column(self.worksheet_mock, **self.args)
        self.worksheet_mock.conditional_format.assert_not_called()


class TestIntegrationTableSectionWriteSupheader:
    def _create_worksheet_with_section_writer_and_write_supheader(
        self, row: int = 1, column: int = 1, num_columns: int = 1, supheader = "Supheader"  # fmt: skip
    ):
        """Creates an xlsxwriter.workbook, an xlsxwriter.worksheet and a
        TableSectionWriter. After using the TableSectionWriter to write a supheader to
        the worksheet, it is then safed to a buffer and loaded with openpyxl. The loaded
        worksheet is returned.

        Note that row and column are specified as one indexed, whereas xlsxwriter is
        zero indexed.
        """
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer._write_supheader(
                worksheet=excel_manager.worksheet,
                row=row - 1,
                column=column - 1,
                num_columns=num_columns,
                supheader=supheader,
                supheader_format={"bold": True},
            )
        return excel_manager.loaded_worksheet

    @pytest.mark.parametrize("row, column", [(1, 2), (2, 1), (5, 10)])
    def test_supheader_written_to_correct_position(self, row, column):
        worksheet = self._create_worksheet_with_section_writer_and_write_supheader(row, column)  # fmt: skip
        assert worksheet.cell(row=row, column=column).value == "Supheader"

    def test_supheader_format_is_applied(self):
        worksheet = self._create_worksheet_with_section_writer_and_write_supheader()
        assert worksheet.cell(row=1, column=1).font.bold == True

    @pytest.mark.parametrize("num_columns", [2, 4, 10])
    def test_supheader_cell_is_merged(self, num_columns):
        worksheet = self._create_worksheet_with_section_writer_and_write_supheader(num_columns=num_columns)  # fmt: skip
        merged_cells = list(worksheet.merged_cells.ranges[0].cells)
        assert merged_cells == [(1, i) for i in range(1, num_columns + 1)]

    def test_no_merge_applied_when_written_with_only_one_column(self):
        worksheet = self._create_worksheet_with_section_writer_and_write_supheader(num_columns=1)  # fmt: skip
        assert len(worksheet.merged_cells.ranges) == 0
        assert worksheet.cell(row=1, column=1).value == "Supheader"

    def test_no_merge_applied_when_no_supheader_specified(self):
        worksheet = self._create_worksheet_with_section_writer_and_write_supheader(num_columns=2, supheader="")  # fmt: skip
        assert len(worksheet.merged_cells.ranges) == 0


class TestIntegrationTableSectionWriteSection:
    def _create_worksheet_with_section_writer_and_write_section(
        self, table_section, write_supheader
    ):
        """Creates an xlsxwriter.workbook, an xlsxwriter.worksheet and a
        TableSectionWriter. After using the TableSectionWriter to write a section to the
        worksheet, it is then safed to a buffer and loaded with openpyxl. The loaded
        worksheet is returned.

        Note that row and column are specified as one indexed, whereas xlsxwriter is
        zero indexed.
        """
        with ExcelWriteReadTestManager() as excel_manager:
            section_writer = writer.TableSectionWriter(excel_manager.workbook)
            section_writer._write_section(
                excel_manager.worksheet,
                table_section,
                start_row=0,
                start_column=0,
                write_supheader=write_supheader,
            )
        return excel_manager.loaded_worksheet

    @pytest.mark.parametrize("write_supheader", [True, False])
    def test_correct_number_of_columns_written(self, table_section, write_supheader):
        worksheet = self._create_worksheet_with_section_writer_and_write_section(
            table_section, write_supheader=write_supheader
        )
        assert len(list(worksheet.columns)) == table_section.data.shape[1]

    @pytest.mark.parametrize("write_supheader, num_headers", [(True, 2), (False, 1)])
    def test_correct_number_of_rows_written(self, table_section, write_supheader, num_headers):  # fmt: skip
        worksheet = self._create_worksheet_with_section_writer_and_write_section(
            table_section, write_supheader=write_supheader
        )
        for column in worksheet.columns:
            assert len(column) == table_section.data.shape[0] + num_headers

    @pytest.mark.parametrize("write_supheader", [True, False])
    def test_supheader_correctly_added_or_ommitted(self, table_section, write_supheader):  # fmt: skip
        table_section.supheader = "A supheader name"
        worksheet = self._create_worksheet_with_section_writer_and_write_section(
            table_section, write_supheader=write_supheader
        )
        # Note that xlsxwriter is zero indexed whereas pyopenxl is one indexed.
        if write_supheader:
            assert worksheet.cell(row=1, column=1).value == table_section.supheader
        else:
            assert worksheet.cell(row=1, column=1).value != table_section.supheader


class TestTableSectionWriteSection:
    @pytest.fixture(autouse=True)
    def _init_section_writer(self):
        self.worksheet_mock = MagicMock(name="worksheet_mock")
        self.section_writer = writer.TableSectionWriter(Workbook())

    @pytest.mark.parametrize("write_supheader", [True, False])
    def test_conditional_format_called_with_correct_arguments(self, table_section, write_supheader):  # fmt: skip
        table_section.section_conditional = {"bold": True}

        self.section_writer._write_section(
            self.worksheet_mock,
            table_section,
            start_row=0,
            start_column=0,
            write_supheader=write_supheader,
        )

        rows, cols = table_section.data.shape
        data_start_row = 2 if write_supheader else 1
        data_end_row = data_start_row + rows - 1
        self.worksheet_mock.conditional_format.assert_called_once_with(
            data_start_row, 0, data_end_row, cols - 1, {"bold": True}
        )

    def test_conditional_format_not_called_when_conditionaL_format_is_empty(self, table_section):  # fmt: skip
        table_section.section_conditional = {}
        self.section_writer._write_section(self.worksheet_mock, table_section, 0, 0, True)  # fmt: skip
        self.worksheet_mock.conditional_format.assert_not_called()

    def test_set_column_not_called_when_hide_section_is_false(self, table_section):
        table_section.hide_section = False
        self.section_writer._write_section(self.worksheet_mock, table_section, 0, 0, True)  # fmt: skip
        self.worksheet_mock.set_column.assert_not_called()

    def test_set_column_called_correctly_when_hide_section_is_false(self, table_section):  # fmt: skip
        table_section.hide_section = True
        self.section_writer._write_section(self.worksheet_mock, table_section, 0, 0, True)  # fmt: skip
        self.worksheet_mock.set_column.assert_called_once_with(
            0, table_section.data.shape[1] - 1, options={"level": 1, "collapsed": True, "hidden": True}  # fmt: skip
        )


class TestTableSectionWriterGetXlsxFormat:
    @pytest.fixture(autouse=True)
    def _init_section_writer(self):
        self.writer = writer.TableSectionWriter(Workbook())
        self.format_description = {"bold": True, "font_color": "#FF0000"}

    def test_xlsx_format_with_correct_properties_is_returned(self):
        xlsx_format = self.writer.get_xlsx_format(self.format_description)
        assert xlsx_format.bold == True and xlsx_format.font_color == "#FF0000"

    def test_multiple_calls_return_same_object(self):
        xlsx_format_1 = self.writer.get_xlsx_format(self.format_description)
        xlsx_format_2 = self.writer.get_xlsx_format(self.format_description)
        assert xlsx_format_1 is xlsx_format_2


class TestHashableFromDictionary:
    @pytest.mark.parametrize(
        "dictionary", [{"A": 1, "C": 2, "B": 1}, {"B": 1, "A": 1, "C": 2}]
    )
    def test_different_key_order_creates_same_hashable(self, dictionary):
        expected_hash = writer._hashable_from_dict({"A": 1, "B": 1, "C": 2})
        assert writer._hashable_from_dict(dictionary) == expected_hash

    def test_different_dict_values_create_different_hashable(self):
        hashable_1 = writer._hashable_from_dict({"A": 1})
        hashable_2 = writer._hashable_from_dict({"A": 2})
        assert hashable_1 != hashable_2
